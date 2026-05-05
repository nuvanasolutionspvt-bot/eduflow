import json
import logging
from io import BytesIO
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from urllib.parse import urlencode
from uuid import uuid4
from zipfile import ZIP_DEFLATED, ZipFile

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth import get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import Group
from django.contrib.auth.views import LoginView, LogoutView
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files import File
from django.db import IntegrityError, transaction
from django.db.models import Count, F, Q, Sum
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.views import View
from django.views.generic import CreateView, DeleteView, FormView, ListView, TemplateView, UpdateView
from PIL import Image, ImageDraw, ImageFont, ImageOps

from .forms import (
    AdminAuthenticationForm,
    BonafideForm,
    ExamForm,
    FeeReceiptForm,
    HallTicketBulkForm,
    HallTicketForm,
    IDCardBulkForm,
    NirgamForm,
    ResultBulkEntryForm,
    ResultForm,
    SchoolRegistrationForm,
    SchoolRoleForm,
    SchoolSettingForm,
    SchoolUserForm,
    StudentBulkImportForm,
    StudentForm,
    StudentSelectForm,
)
from .models import (
    DocumentCounter,
    DocumentTemplateSetting,
    Exam,
    ExamSubject,
    FeeReceipt,
    Result,
    School,
    SchoolRole,
    SchoolSetting,
    SchoolSubscription,
    SubscriptionPayment,
    Student,
    UserProfile,
)
from .translation import TranslationServiceError, translate_text


logger = logging.getLogger(__name__)


def _resolve_school_for_user(user):
    if user is None or not user.is_authenticated:
        return None

    profile = getattr(user, 'tenant_profile', None)
    if profile is not None and profile.school is not None:
        return profile.school

    profile = UserProfile.objects.filter(user=user).select_related('school').first()
    if profile is not None and profile.school is not None:
        return profile.school

    return getattr(user, 'school', None)


def _require_school(request):
    school = getattr(request, 'school', None)
    if school is None:
        school = _resolve_school_for_user(getattr(request, 'user', None))
        if school is not None:
            request.school = school
    if school is None:
        raise PermissionDenied('The authenticated user is not assigned to a school.')
    return school


def _student_queryset(request):
    return Student.objects.filter(school=_require_school(request))


def _exam_queryset(request):
    return Exam.objects.filter(school=_require_school(request))


def _fee_receipt_queryset(request):
    return FeeReceipt.objects.filter(school=_require_school(request)).select_related('student')


def _result_queryset(request):
    return Result.objects.filter(school=_require_school(request)).select_related('student', 'exam', 'subject')


def _school_role_queryset(request):
    return SchoolRole.objects.filter(school=_require_school(request)).select_related('group')


def _school_user_queryset(request):
    school = _require_school(request)
    return get_user_model().objects.filter(tenant_profile__school=school).select_related('tenant_profile')


def _school_setting(request):
    school = _require_school(request)
    return SchoolSetting.objects.filter(school=school).first()


def _document_template_settings(request, document_type):
    school = _require_school(request)
    return DocumentTemplateSetting.objects.filter(school=school, document_type=document_type).first()


def _clamp_int(value, default, minimum, maximum):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return max(minimum, min(parsed, maximum))


def _clean_hex_color(value, default):
    value = (value or '').strip()
    if len(value) == 7 and value.startswith('#'):
        return value
    return default


def _school_subscription(school):
    if school is None:
        return None

    created_at = getattr(school, 'created_at', None) or timezone.now()
    trial_started_on = timezone.localtime(created_at).date() if timezone.is_aware(created_at) else created_at.date()
    subscription, _ = SchoolSubscription.objects.get_or_create(
        school=school,
        defaults={
            'trial_started_on': trial_started_on,
            'trial_expires_on': trial_started_on + timedelta(days=3),
            'annual_price': 5999,
        },
    )
    return subscription


def _subscription_redirect(request, reason='download'):
    subscribe_url = reverse('subscription')
    query = urlencode(
        {
            'next': request.get_full_path(),
            'reason': reason,
        }
    )
    messages.warning(
        request,
        'Your free download has already been used. Subscribe to continue downloading.',
    )
    return redirect(f'{subscribe_url}?{query}')


def _authorize_download(request, reason='download'):
    school = _require_school(request)
    subscription = _school_subscription(school)

    if subscription.annual_plan_is_active:
        return None

    with transaction.atomic():
        locked = SchoolSubscription.objects.select_for_update().get(pk=subscription.pk)
        if locked.annual_plan_is_active:
            return None
        if locked.remaining_free_downloads > 0:
            locked.free_downloads_used += 1
            locked.save(update_fields=['free_downloads_used', 'updated_at'])
            return None

    return _subscription_redirect(request, reason=reason)


def _razorpay_is_configured():
    return bool(settings.RAZORPAY_KEY_ID and settings.RAZORPAY_KEY_SECRET)


def _razorpay_client():
    if not _razorpay_is_configured():
        raise RuntimeError('Razorpay credentials are not configured.')

    try:
        import razorpay
    except ModuleNotFoundError as exc:
        raise RuntimeError('Razorpay SDK is not installed.') from exc

    client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))
    client.set_app_details({"title": "EduFlow", "version": "1.0"})
    return client


class SchoolAccessMixin(LoginRequiredMixin):
    required_permissions = ()

    def dispatch(self, request, *args, **kwargs):
        school = _resolve_school_for_user(getattr(request, 'user', None))
        if school is None:
            messages.error(request, 'Your account is not assigned to a school yet.')
            return redirect('school_register')
        request.school = school
        if not self._has_required_permissions(request.user):
            raise PermissionDenied('You do not have permission to access this module.')
        return super().dispatch(request, *args, **kwargs)

    def _has_required_permissions(self, user):
        if user.is_superuser or user.is_staff:
            return True
        profile = getattr(user, 'tenant_profile', None)
        if profile is not None and profile.is_school_admin:
            return True
        required = self.required_permissions
        if isinstance(required, str):
            required = (required,)
        if not required:
            return True
        return user.has_perms(required)


class SchoolScopedModelFormMixin(SchoolAccessMixin):
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['school'] = self.request.school
        return kwargs


class SchoolScopedStudentObjectMixin(SchoolAccessMixin):
    def get_queryset(self):
        return _student_queryset(self.request)


class WorkspaceEntryView(View):
    def get(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')

        school = _resolve_school_for_user(request.user)
        if school is None:
            messages.error(request, 'Your account is not assigned to a school yet.')
            return redirect('school_register')

        request.school = school
        return redirect('dashboard')


class AdminLoginView(LoginView):
    template_name = 'registration/login.html'
    authentication_form = AdminAuthenticationForm

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            if _resolve_school_for_user(request.user) is None:
                return redirect('school_register')
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['has_admin_user'] = get_user_model().objects.exists()
        return context


class SchoolRegistrationView(FormView):
    template_name = 'registration/register.html'
    form_class = SchoolRegistrationForm
    success_url = reverse_lazy('dashboard')

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            if _resolve_school_for_user(request.user) is not None:
                return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        with transaction.atomic():
            school = School.objects.create(
                name=form.cleaned_data['school_name'],
                address=form.cleaned_data['school_address'],
                logo=form.cleaned_data.get('school_logo'),
            )
            user = form.save(commit=False)
            user.email = form.cleaned_data['admin_email']
            user.is_staff = True
            user.is_active = True
            user.save()
            UserProfile.objects.create(user=user, school=school)
            SchoolSetting.objects.create(
                school=school,
                school_name=school.name,
                address=school.address,
                logo=school.logo,
                principal_name=form.cleaned_data['principal_name'],
            )

        login(self.request, user)
        messages.success(self.request, 'School workspace created successfully.')
        return super().form_valid(form)


class AdminLogoutView(LogoutView):
    pass


class DashboardView(SchoolAccessMixin, TemplateView):
    template_name = 'core/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        students = _student_queryset(self.request)
        fee_receipts = _fee_receipt_queryset(self.request)
        subscription = _school_subscription(self.request.school)
        context['total_students'] = students.count()
        gender_totals = {
            item['gender']: item['total']
            for item in students.values('gender').annotate(total=Count('id'))
        }
        male_students = gender_totals.get('male', 0)
        female_students = gender_totals.get('female', 0)
        gender_total = male_students + female_students
        male_percent = (male_students / gender_total * 100) if gender_total else 0
        context['gender_chart'] = {
            'male': male_students,
            'female': female_students,
            'total': gender_total,
            'male_percent': male_percent,
            'female_percent': 100 - male_percent if gender_total else 0,
            'male_angle': male_percent * 3.6,
        }
        context['class_counts'] = (
            students.values('student_class').annotate(total=Count('id')).order_by('student_class')
        )
        fee_years = list(
            fee_receipts
            .values_list('receipt_date__year', flat=True)
            .distinct()
            .order_by('-receipt_date__year')
        )
        selected_fee_year = self.request.GET.get('fee_year')
        try:
            selected_fee_year = int(selected_fee_year)
        except (TypeError, ValueError):
            selected_fee_year = fee_years[0] if fee_years else timezone.localdate().year
        if fee_years and selected_fee_year not in fee_years:
            selected_fee_year = fee_years[0]

        fee_total_expression = (
            F('tuition_fee')
            + F('library_fee')
            + F('lab_fee')
            + F('sports_fee')
            + F('other_fee')
            + F('donation')
        )
        selected_fee_summary = fee_receipts.filter(receipt_date__year=selected_fee_year).aggregate(
            total_fees=Sum(fee_total_expression),
            collected_fees=Sum('paid_amount'),
            remaining_fees=Sum(fee_total_expression - F('paid_amount')),
        )
        context['fee_years'] = fee_years
        context['selected_fee_year'] = selected_fee_year
        context['selected_fee_summary'] = {
            'total_fees': selected_fee_summary['total_fees'] or 0,
            'collected_fees': selected_fee_summary['collected_fees'] or 0,
            'remaining_fees': selected_fee_summary['remaining_fees'] or 0,
        }
        context['document_counts'] = _document_count_totals(self.request.school)
        context['subscription'] = subscription
        return context


class StudentListView(SchoolAccessMixin, ListView):
    required_permissions = 'core.view_student'
    model = Student
    template_name = 'core/student_list.html'
    context_object_name = 'students'
    paginate_by = 10

    def get_queryset(self):
        queryset = _student_queryset(self.request)
        query = self.request.GET.get('q', '').strip()
        class_filter = self.request.GET.get('class', '').strip()
        section_filter = self.request.GET.get('section', '').strip()

        if query:
            queryset = queryset.filter(
                Q(name__icontains=query)
                | Q(admission_no__icontains=query)
                | Q(roll_no__icontains=query)
            )
        if class_filter:
            queryset = queryset.filter(student_class=class_filter)
        if section_filter:
            queryset = queryset.filter(section=section_filter)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['student_classes'] = Student.CLASS_CHOICES
        context['sections'] = Student.SECTION_CHOICES
        context['filters'] = {
            'q': self.request.GET.get('q', ''),
            'class': self.request.GET.get('class', ''),
            'section': self.request.GET.get('section', ''),
        }
        return context


class StudentCreateView(SchoolScopedModelFormMixin, CreateView):
    required_permissions = 'core.add_student'
    model = Student
    form_class = StudentForm
    template_name = 'core/student_form.html'
    success_url = reverse_lazy('student_list')

    def form_valid(self, form):
        messages.success(self.request, 'Student added successfully.')
        return super().form_valid(form)


class StudentUpdateView(SchoolScopedStudentObjectMixin, SchoolScopedModelFormMixin, UpdateView):
    required_permissions = 'core.change_student'
    model = Student
    form_class = StudentForm
    template_name = 'core/student_form.html'
    success_url = reverse_lazy('student_list')

    def form_valid(self, form):
        messages.success(self.request, 'Student updated successfully.')
        return super().form_valid(form)


class StudentDeleteView(SchoolScopedStudentObjectMixin, SchoolAccessMixin, DeleteView):
    required_permissions = 'core.delete_student'
    model = Student
    success_url = reverse_lazy('student_list')

    def post(self, request, *args, **kwargs):
        messages.success(request, 'Student deleted successfully.')
        return super().post(request, *args, **kwargs)


class StudentTransliterateView(SchoolAccessMixin, View):
    def _extract_text(self, request):
        if request.method == 'GET':
            return (request.GET.get('text') or '').strip(), None
        if request.method != 'POST':
            return '', JsonResponse({'error': 'Method not allowed.'}, status=405)

        try:
            payload = json.loads(request.body.decode('utf-8') or '{}')
        except json.JSONDecodeError:
            return '', JsonResponse({'error': 'Invalid JSON payload.'}, status=400)
        return (payload.get('text') or '').strip(), None

    def get(self, request):
        return self._translate(request)

    def post(self, request):
        return self._translate(request)

    def _translate(self, request):
        text, error_response = self._extract_text(request)
        if error_response is not None:
            return error_response
        if not text:
            return JsonResponse({'text': ''})

        try:
            translated_text = translate_text(text, fail_silently=False)
        except TranslationServiceError as exc:
            return JsonResponse({'error': str(exc)}, status=503)

        return JsonResponse({'text': translated_text})


class StudentBulkImportView(SchoolAccessMixin, FormView):
    required_permissions = 'core.add_student'
    form_class = StudentBulkImportForm
    template_name = 'core/student_bulk_import.html'
    success_url = reverse_lazy('student_list')

    SHORT_HEADERS = [
        'name',
        'father_name',
        'dob',
        'gender',
        'student_class',
        'section',
        'roll_no',
        'admission_no',
        'address',
        'mobile',
        'status',
    ]
    SHORT_HEADERS_WITH_PHOTO = SHORT_HEADERS[:-1] + ['photo', 'status']
    REQUIRED_HEADERS = [
        'name',
        'father_name',
        'mother_name',
        'religion',
        'caste',
        'mother_tongue',
        'father_occupation',
        'dob',
        'gender',
        'birth_place',
        'taluka',
        'district',
        'student_class',
        'section',
        'roll_no',
        'admission_no',
        'register_page_no',
        'previous_school_name',
        'previous_school_class',
        'address',
        'mobile',
        'status',
    ]
    REQUIRED_HEADERS_WITH_PHOTO = REQUIRED_HEADERS[:-1] + ['photo', 'status']
    DATE_INPUT_FORMATS = (
        '%d/%m/%Y',
        '%d-%m-%Y',
        '%d.%m.%Y',
        '%Y/%m/%d',
        '%Y-%m-%d',
        '%Y.%m.%d',
        '%m/%d/%Y',
        '%m-%d-%Y',
        '%m.%d.%Y',
    )

    def form_valid(self, form):
        try:
            from openpyxl import load_workbook
        except Exception:
            form.add_error(
                'excel_file',
                'Excel support is not installed. Please install openpyxl and try again.',
            )
            return self.form_invalid(form)

        excel_file = form.cleaned_data['excel_file']
        try:
            workbook = load_workbook(excel_file, data_only=True)
        except Exception:
            form.add_error('excel_file', 'Could not read Excel file. Please use a valid .xlsx file.')
            return self.form_invalid(form)

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            form.add_error('excel_file', 'Excel file is empty.')
            return self.form_invalid(form)

        header = [str(value).strip() if value is not None else '' for value in rows[0]]
        accepted_headers = (
            self.SHORT_HEADERS,
            self.SHORT_HEADERS_WITH_PHOTO,
            self.REQUIRED_HEADERS,
            self.REQUIRED_HEADERS_WITH_PHOTO,
        )
        active_headers = next((candidate for candidate in accepted_headers if header == candidate), None)
        if active_headers is None:
            form.add_error(
                'excel_file',
                (
                    'Invalid format. Please download and use the provided template. '
                    'Accepted formats: short or full template, each with or without optional photo column.'
                ),
            )
            return self.form_invalid(form)

        created_count = 0
        failed_rows = []
        for index, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue
            row_data = dict(zip(active_headers, row))
            payload = {}
            for key, value in row_data.items():
                if key == 'photo':
                    continue
                payload[key] = str(value).strip() if value is not None else ''

            # Accept Excel date cells and ISO strings for DOB.
            payload['dob'] = self._normalize_excel_dob(row_data.get('dob'))

            photo_raw = row_data.get('photo')
            photo_value = str(photo_raw).strip() if photo_raw is not None else ''
            photo_path = None
            if photo_value:
                photo_path = self._resolve_photo_path(photo_value)
                if not photo_path:
                    failed_rows.append(
                        (
                            f'Row {index}: photo file not found for "{photo_value}". '
                            'Use absolute path, project-relative path, or media-relative path.'
                        )
                    )
                    continue

            student_form = StudentForm(payload, school=self.request.school)
            if student_form.is_valid():
                try:
                    with transaction.atomic():
                        student = student_form.save()
                        if photo_path:
                            with open(photo_path, 'rb') as photo_file:
                                student.photo.save(photo_path.name, File(photo_file), save=False)
                            student.save()
                    created_count += 1
                except IntegrityError as exc:
                    failed_rows.append(f'Row {index}: {exc}')
                except Exception as exc:
                    failed_rows.append(f'Row {index}: photo upload failed: {exc}')
            else:
                error_summary = '; '.join(
                    [f'{field}: {", ".join(errors)}' for field, errors in student_form.errors.items()]
                )
                failed_rows.append(f'Row {index}: {error_summary}')

        if failed_rows:
            for error in failed_rows[:10]:
                messages.error(self.request, error)
            if len(failed_rows) > 10:
                messages.error(self.request, f'...and {len(failed_rows) - 10} more row errors.')

        if created_count:
            messages.success(self.request, f'{created_count} students imported successfully.')
        elif not failed_rows:
            messages.info(self.request, 'No student rows found to import.')

        return super().form_valid(form)

    def _resolve_photo_path(self, photo_value):
        given_path = Path(photo_value)
        candidates = []

        if given_path.is_absolute():
            candidates.append(given_path)
        else:
            candidates.append(Path(settings.BASE_DIR) / given_path)
            candidates.append(Path(settings.MEDIA_ROOT) / given_path)
            candidates.append(Path(settings.MEDIA_ROOT) / 'students' / 'photos' / given_path)

        for candidate in candidates:
            if candidate.exists() and candidate.is_file():
                return candidate
        return None

    def _normalize_excel_dob(self, dob_value):
        if isinstance(dob_value, datetime):
            return dob_value.date().isoformat()
        if isinstance(dob_value, date):
            return dob_value.isoformat()
        if dob_value is None:
            return ''

        if isinstance(dob_value, (int, float)):
            try:
                from openpyxl.utils.datetime import from_excel

                normalized = from_excel(dob_value)
                if isinstance(normalized, datetime):
                    return normalized.date().isoformat()
                if isinstance(normalized, date):
                    return normalized.isoformat()
            except Exception:
                return str(dob_value).strip()

        raw_value = str(dob_value).strip()
        if not raw_value:
            return ''

        parsed_value = parse_date(raw_value)
        if parsed_value:
            return parsed_value.isoformat()

        for fmt in self.DATE_INPUT_FORMATS:
            try:
                return datetime.strptime(raw_value, fmt).date().isoformat()
            except ValueError:
                continue

        return raw_value


class StudentImportTemplateView(SchoolAccessMixin, View):
    HEADERS = [
        'name',
        'father_name',
        'mother_name',
        'religion',
        'caste',
        'mother_tongue',
        'father_occupation',
        'dob',
        'gender',
        'birth_place',
        'taluka',
        'district',
        'student_class',
        'section',
        'roll_no',
        'admission_no',
        'register_page_no',
        'previous_school_name',
        'previous_school_class',
        'address',
        'mobile',
        'photo',
        'status',
    ]

    SAMPLE_ROW = [
        'Aarav Sharma',
        'Rakesh Sharma',
        'Sunita Sharma',
        'Hindu',
        'Open',
        'Marathi',
        'Farmer',
        '2012-06-15',
        'male',
        'Nanded',
        'Nanded',
        'Nanded',
        '8',
        'A',
        '12',
        'ADM-2026-001',
        '45',
        'ZP Primary School',
        '7',
        '123 Main Road, City',
        '9876543210',
        'students/photos/aarav.jpg',
        'active',
    ]

    def get(self, request):
        try:
            from openpyxl import Workbook
        except Exception:
            messages.error(request, 'Excel support is not installed. Please install openpyxl first.')
            return redirect('student_bulk_import')

        blocked_response = _authorize_download(request, reason='student-import-template')
        if blocked_response is not None:
            return blocked_response

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Students'
        sheet.append(self.HEADERS)
        sheet.append(self.SAMPLE_ROW)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
        workbook.save(response)
        return response


class SchoolSettingUpdateView(SchoolAccessMixin, UpdateView):
    required_permissions = 'core.change_schoolsetting'
    model = SchoolSetting
    form_class = SchoolSettingForm
    template_name = 'core/school_setting_form.html'
    success_url = reverse_lazy('school_settings')

    def get_object(self, queryset=None):
        school = self.request.school
        setting, _ = SchoolSetting.objects.get_or_create(
            school=school,
            defaults={
                'school_name': school.name,
                'address': school.address or 'School address',
                'principal_name': 'Principal Name',
            },
        )
        return setting

    def form_valid(self, form):
        messages.success(self.request, 'School settings updated successfully.')
        return super().form_valid(form)


class ExamListView(SchoolAccessMixin, ListView):
    required_permissions = 'core.view_exam'
    model = Exam
    template_name = 'core/exam_list.html'
    context_object_name = 'exams'

    def get_queryset(self):
        return _exam_queryset(self.request)


class ExamCreateView(SchoolAccessMixin, View):
    required_permissions = 'core.add_exam'
    template_name = 'core/exam_form.html'

    def get(self, request):
        form = ExamForm()
        return render(request, self.template_name, {'form': form, 'title': 'Add Exam'})

    def post(self, request):
        form = ExamForm(request.POST)
        if form.is_valid():
            exam = form.save(commit=False)
            exam.school = request.school
            exam.exam_date = date.today()
            exam.save()
            messages.success(request, 'Exam created successfully.')
            return redirect('exam_list')
        return render(request, self.template_name, {'form': form, 'title': 'Add Exam'})


class ExamUpdateView(SchoolAccessMixin, View):
    required_permissions = 'core.change_exam'
    template_name = 'core/exam_form.html'

    def get(self, request, pk):
        exam = get_object_or_404(_exam_queryset(request), pk=pk)
        form = ExamForm(instance=exam)
        return render(
            request,
            self.template_name,
            {'form': form, 'exam': exam, 'title': 'Edit Exam'},
        )

    def post(self, request, pk):
        exam = get_object_or_404(_exam_queryset(request), pk=pk)
        form = ExamForm(request.POST, instance=exam)
        if form.is_valid():
            updated_exam = form.save(commit=False)
            updated_exam.school = request.school
            updated_exam.save()
            messages.success(request, 'Exam updated successfully.')
            return redirect('exam_list')
        return render(
            request,
            self.template_name,
            {'form': form, 'exam': exam, 'title': 'Edit Exam'},
        )


class ExamDeleteView(SchoolAccessMixin, View):
    required_permissions = 'core.delete_exam'
    def post(self, request, pk):
        exam = get_object_or_404(_exam_queryset(request), pk=pk)
        exam.delete()
        messages.success(request, 'Exam deleted successfully.')
        return redirect('exam_list')


class FeeReceiptListView(SchoolAccessMixin, ListView):
    required_permissions = 'core.view_feereceipt'
    model = FeeReceipt
    template_name = 'core/fee_receipt_list.html'
    context_object_name = 'receipts'
    paginate_by = 10

    def get_queryset(self):
        queryset = _fee_receipt_queryset(self.request)
        query = self.request.GET.get('q', '').strip()
        class_filter = self.request.GET.get('class', '').strip()

        if query:
            queryset = queryset.filter(
                Q(receipt_no__icontains=query)
                | Q(student__name__icontains=query)
                | Q(student__admission_no__icontains=query)
            )
        if class_filter:
            queryset = queryset.filter(student__student_class=class_filter)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['student_classes'] = Student.CLASS_CHOICES
        context['filters'] = {
            'q': self.request.GET.get('q', ''),
            'class': self.request.GET.get('class', ''),
        }
        return context


class FeeReceiptCreateView(SchoolScopedModelFormMixin, CreateView):
    required_permissions = 'core.add_feereceipt'
    model = FeeReceipt
    form_class = FeeReceiptForm
    template_name = 'core/fee_receipt_form.html'

    def form_valid(self, form):
        form.instance.receipt_no = _next_serial(self.request.school, 'fee-receipt')
        messages.success(self.request, 'Fee receipt created successfully.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('fee_receipt_detail', kwargs={'pk': self.object.pk})


class FeeReceiptUpdateView(SchoolScopedModelFormMixin, UpdateView):
    required_permissions = 'core.change_feereceipt'
    model = FeeReceipt
    form_class = FeeReceiptForm
    template_name = 'core/fee_receipt_form.html'

    def get_queryset(self):
        return _fee_receipt_queryset(self.request)

    def form_valid(self, form):
        messages.success(self.request, 'Fee receipt updated successfully.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('fee_receipt_detail', kwargs={'pk': self.object.pk})


class FeeReceiptDetailView(SchoolAccessMixin, View):
    required_permissions = 'core.view_feereceipt'
    template_name = 'core/fee_receipt_detail.html'

    def get(self, request, pk):
        receipt = get_object_or_404(_fee_receipt_queryset(request), pk=pk)
        return render(request, self.template_name, {'receipt': receipt})


class FeeReceiptDeleteView(SchoolAccessMixin, View):
    required_permissions = 'core.delete_feereceipt'
    def post(self, request, pk):
        receipt = get_object_or_404(_fee_receipt_queryset(request), pk=pk)
        receipt.delete()
        messages.success(request, 'Fee receipt deleted successfully.')
        return redirect('fee_receipt_list')


class ResultListView(SchoolAccessMixin, ListView):
    required_permissions = 'core.view_result'
    model = Result
    template_name = 'core/result_list.html'
    context_object_name = 'report_cards'
    paginate_by = 20

    def get_queryset(self):
        queryset = _result_queryset(self.request)
        query = self.request.GET.get('q', '').strip()
        exam_filter = self.request.GET.get('exam', '').strip()
        class_filter = self.request.GET.get('class', '').strip()

        if query:
            queryset = queryset.filter(
                Q(student__name__icontains=query)
                | Q(student__admission_no__icontains=query)
                | Q(student__roll_no__icontains=query)
                | Q(subject__subject_name__icontains=query)
            )
        if exam_filter:
            queryset = queryset.filter(exam_id=exam_filter)
        if class_filter:
            queryset = queryset.filter(student__student_class=class_filter)
        return queryset.order_by('student__student_class', 'student__section', 'student__roll_no', 'student__name')

    def paginate_queryset(self, queryset, page_size):
        return super().paginate_queryset(list(self._build_report_cards(queryset)), page_size)

    def _build_report_cards(self, queryset):
        grouped = {}
        for result in queryset:
            key = result.student_id
            if key not in grouped:
                grouped[key] = {
                    'student': result.student,
                    'first_exam': result.exam,
                    'exam_names': [],
                    'subject_names': set(),
                    'total_obtained': Decimal('0'),
                    'total_marks': Decimal('0'),
                    'result_count': 0,
                }
            item = grouped[key]
            if result.exam.name not in item['exam_names']:
                item['exam_names'].append(result.exam.name)
            item['subject_names'].add(result.subject.subject_name)
            item['total_obtained'] += result.marks_obtained
            item['total_marks'] += result.max_marks
            item['result_count'] += 1

        for item in grouped.values():
            item['subject_count'] = len(item['subject_names'])
            item['percentage'] = round((item['total_obtained'] / item['total_marks']) * 100, 2) if item['total_marks'] else 0
            item['status'] = 'Pass' if item['percentage'] >= 35 else 'Fail'
        return grouped.values()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['student_classes'] = Student.CLASS_CHOICES
        context['exams'] = _exam_queryset(self.request).order_by('-id')
        context['filters'] = {
            'q': self.request.GET.get('q', ''),
            'exam': self.request.GET.get('exam', ''),
            'class': self.request.GET.get('class', ''),
        }
        return context


class ResultCreateView(SchoolAccessMixin, View):
    required_permissions = 'core.add_result'
    template_name = 'core/result_bulk_form.html'

    def get(self, request):
        form = ResultBulkEntryForm(school=request.school)
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = ResultBulkEntryForm(request.POST, school=request.school)
        rows = self._posted_rows(request)
        row_errors = []

        if form.is_valid():
            saved_count = self._save_rows(form.cleaned_data, rows, row_errors)
            if saved_count and not row_errors:
                messages.success(request, f'{saved_count} result entries saved successfully.')
                return redirect('result_list')
            if saved_count:
                messages.warning(request, f'{saved_count} result entries saved. Some rows need correction.')
        return render(
            request,
            self.template_name,
            {
                'form': form,
                'posted_rows': rows,
                'row_errors': row_errors,
            },
        )

    def _posted_rows(self, request):
        subjects = request.POST.getlist('subject_name[]')
        first_marks = request.POST.getlist('first_marks[]')
        first_max = request.POST.getlist('first_max[]')
        second_marks = request.POST.getlist('second_marks[]')
        second_max = request.POST.getlist('second_max[]')
        remarks = request.POST.getlist('remarks[]')
        row_count = max(
            len(subjects),
            len(first_marks),
            len(first_max),
            len(second_marks),
            len(second_max),
            len(remarks),
            1,
        )
        rows = []
        for index in range(row_count):
            rows.append(
                {
                    'subject_name': subjects[index].strip() if index < len(subjects) else '',
                    'first_marks': first_marks[index].strip() if index < len(first_marks) else '',
                    'first_max': first_max[index].strip() if index < len(first_max) else '100',
                    'second_marks': second_marks[index].strip() if index < len(second_marks) else '',
                    'second_max': second_max[index].strip() if index < len(second_max) else '100',
                    'remarks': remarks[index].strip() if index < len(remarks) else '',
                }
            )
        return rows

    def _decimal_value(self, value, default=None):
        if value in (None, ''):
            return default
        try:
            parsed = Decimal(str(value))
        except Exception:
            return None
        return parsed if parsed >= 0 else None

    def _save_rows(self, cleaned_data, rows, row_errors):
        student = cleaned_data['student']
        first_term = cleaned_data['first_term']
        second_term = cleaned_data.get('second_term')
        saved_count = 0

        with transaction.atomic():
            for index, row in enumerate(rows, start=1):
                subject_name = row['subject_name']
                if not subject_name:
                    if any(row.get(key) for key in ['first_marks', 'second_marks', 'remarks']):
                        row_errors.append(f'Row {index}: subject is required.')
                    continue

                first_marks = self._decimal_value(row['first_marks'])
                first_max = self._decimal_value(row['first_max'], Decimal('100'))
                second_marks = self._decimal_value(row['second_marks'])
                second_max = self._decimal_value(row['second_max'], Decimal('100'))

                if first_marks is None and second_marks is None:
                    row_errors.append(f'Row {index}: enter first term or second term marks.')
                    continue
                if first_max is None or first_max <= 0:
                    row_errors.append(f'Row {index}: first term max marks must be greater than zero.')
                    continue
                if second_marks is not None and second_term is None:
                    row_errors.append(f'Row {index}: select second term before entering second term marks.')
                    continue
                if second_marks is not None and (second_max is None or second_max <= 0):
                    row_errors.append(f'Row {index}: second term max marks must be greater than zero.')
                    continue
                if first_marks is not None and first_marks > first_max:
                    row_errors.append(f'Row {index}: first term marks cannot exceed max marks.')
                    continue
                if second_marks is not None and second_marks > second_max:
                    row_errors.append(f'Row {index}: second term marks cannot exceed max marks.')
                    continue

                saved_count += self._upsert_result(
                    student=student,
                    exam=first_term,
                    subject_name=subject_name,
                    marks_obtained=first_marks,
                    max_marks=first_max,
                    remarks=row['remarks'],
                )
                if second_term is not None:
                    saved_count += self._upsert_result(
                        student=student,
                        exam=second_term,
                        subject_name=subject_name,
                        marks_obtained=second_marks,
                        max_marks=second_max,
                        remarks=row['remarks'],
                    )
        return saved_count

    def _upsert_result(self, student, exam, subject_name, marks_obtained, max_marks, remarks):
        if marks_obtained is None:
            return 0
        subject = ExamSubject.objects.filter(
            school=self.request.school,
            exam=exam,
            subject_name__iexact=subject_name,
        ).first()
        if subject is None:
            subject = ExamSubject.objects.create(
                school=self.request.school,
                exam=exam,
                subject_name=subject_name,
                exam_date=exam.exam_date,
            )
        Result.objects.update_or_create(
            school=self.request.school,
            student=student,
            exam=exam,
            subject=subject,
            defaults={
                'marks_obtained': marks_obtained,
                'max_marks': max_marks,
                'remarks': remarks,
            },
        )
        return 1


class ResultUpdateView(SchoolScopedModelFormMixin, UpdateView):
    required_permissions = 'core.change_result'
    model = Result
    form_class = ResultForm
    template_name = 'core/result_form.html'
    success_url = reverse_lazy('result_list')

    def get_queryset(self):
        return _result_queryset(self.request)

    def form_valid(self, form):
        messages.success(self.request, 'Result updated successfully.')
        return super().form_valid(form)


class ResultDeleteView(SchoolAccessMixin, View):
    required_permissions = 'core.delete_result'
    def post(self, request, pk):
        result = get_object_or_404(_result_queryset(request), pk=pk)
        result.delete()
        messages.success(request, 'Result deleted successfully.')
        return redirect('result_list')


class ResultReportCardView(SchoolAccessMixin, View):
    required_permissions = 'core.view_result'
    template_name = 'core/result_report_card.html'

    def get(self, request, student_id, exam_id):
        student = get_object_or_404(_student_queryset(request), pk=student_id)
        first_term = get_object_or_404(_exam_queryset(request), pk=exam_id)
        second_term_id = request.GET.get('second_term')
        second_term = None
        if second_term_id:
            second_term = get_object_or_404(_exam_queryset(request), pk=second_term_id)
        else:
            second_term = (
                _result_queryset(request)
                .filter(student=student)
                .exclude(exam=first_term)
                .order_by('exam__exam_date', 'exam__name')
                .values_list('exam', flat=True)
                .first()
            )
            if second_term is not None:
                second_term = get_object_or_404(_exam_queryset(request), pk=second_term)

        term_results = list(
            _result_queryset(request)
            .filter(student=student, exam__in=[term for term in [first_term, second_term] if term is not None])
            .order_by('subject__subject_name', 'exam__exam_date', 'exam__name')
        )
        rows_by_subject = {}
        for result in term_results:
            subject_key = result.subject.subject_name.strip().lower()
            if subject_key not in rows_by_subject:
                rows_by_subject[subject_key] = {
                    'subject_name': result.subject.subject_name,
                    'first': None,
                    'second': None,
                }
            if result.exam_id == first_term.id:
                rows_by_subject[subject_key]['first'] = result
            elif second_term is not None and result.exam_id == second_term.id:
                rows_by_subject[subject_key]['second'] = result

        report_rows = sorted(rows_by_subject.values(), key=lambda row: row['subject_name'].lower())
        for row in report_rows:
            row['total_obtained'] = sum(
                (
                    result.marks_obtained
                    for result in [row['first'], row['second']]
                    if result is not None
                ),
                Decimal('0'),
            )
            row['total_marks'] = sum(
                (
                    result.max_marks
                    for result in [row['first'], row['second']]
                    if result is not None
                ),
                Decimal('0'),
            )
        first_total_obtained = sum((row['first'].marks_obtained for row in report_rows if row['first']), Decimal('0'))
        first_total_marks = sum((row['first'].max_marks for row in report_rows if row['first']), Decimal('0'))
        second_total_obtained = sum((row['second'].marks_obtained for row in report_rows if row['second']), Decimal('0'))
        second_total_marks = sum((row['second'].max_marks for row in report_rows if row['second']), Decimal('0'))
        total_obtained = first_total_obtained + second_total_obtained
        total_marks = first_total_marks + second_total_marks
        percentage = round((total_obtained / total_marks) * 100, 2) if total_marks else 0
        status = 'Pass' if percentage >= 35 else 'Fail'
        template_setting = _document_template_settings(request, 'report-card')
        school_setting = _school_setting(request)

        return render(
            request,
            self.template_name,
            {
                'student': student,
                'first_term': first_term,
                'second_term': second_term,
                'report_rows': report_rows,
                'first_total_obtained': first_total_obtained,
                'first_total_marks': first_total_marks,
                'second_total_obtained': second_total_obtained,
                'second_total_marks': second_total_marks,
                'total_obtained': total_obtained,
                'total_marks': total_marks,
                'percentage': percentage,
                'status': status,
                'template_setting': template_setting,
                'school_setting': school_setting,
            },
        )


class RoleControlView(SchoolAccessMixin, TemplateView):
    required_permissions = 'core.view_schoolrole'
    template_name = 'core/role_control.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['roles'] = _school_role_queryset(self.request).prefetch_related('group__permissions')
        users = _school_user_queryset(self.request).prefetch_related('groups')
        school_roles = list(_school_role_queryset(self.request))
        role_by_group_id = {role.group_id: role for role in school_roles}
        context['users'] = [
            {
                'user': user,
                'roles': [role_by_group_id[group.id] for group in user.groups.all() if group.id in role_by_group_id],
            }
            for user in users
        ]
        return context


class SchoolRoleCreateView(SchoolAccessMixin, View):
    required_permissions = 'core.add_schoolrole'
    template_name = 'core/role_form.html'

    def get(self, request):
        return render(request, self.template_name, {'form': SchoolRoleForm(school=request.school), 'title': 'Add Role'})

    def post(self, request):
        form = SchoolRoleForm(request.POST, school=request.school)
        if form.is_valid():
            role_name = form.cleaned_data['name'].strip()
            group = Group.objects.create(name=f'school-{request.school.id}-{role_name}')
            group.permissions.set(form.selected_permissions())
            SchoolRole.objects.create(school=request.school, group=group, name=role_name)
            messages.success(request, 'Role created successfully.')
            return redirect('role_control')
        return render(request, self.template_name, {'form': form, 'title': 'Add Role'})


class SchoolRoleUpdateView(SchoolAccessMixin, View):
    required_permissions = 'core.change_schoolrole'
    template_name = 'core/role_form.html'

    def get(self, request, pk):
        role = get_object_or_404(_school_role_queryset(request), pk=pk)
        return render(request, self.template_name, {'form': SchoolRoleForm(instance=role, school=request.school), 'role': role, 'title': 'Edit Role'})

    def post(self, request, pk):
        role = get_object_or_404(_school_role_queryset(request), pk=pk)
        form = SchoolRoleForm(request.POST, instance=role, school=request.school)
        if form.is_valid():
            updated_role = form.save(commit=False)
            updated_role.school = request.school
            updated_role.group = role.group
            updated_role.save()
            role.group.name = f'school-{request.school.id}-{updated_role.name}'
            role.group.save(update_fields=['name'])
            role.group.permissions.set(form.selected_permissions())
            messages.success(request, 'Role updated successfully.')
            return redirect('role_control')
        return render(request, self.template_name, {'form': form, 'role': role, 'title': 'Edit Role'})


class SchoolRoleDeleteView(SchoolAccessMixin, View):
    required_permissions = 'core.delete_schoolrole'

    def post(self, request, pk):
        role = get_object_or_404(_school_role_queryset(request), pk=pk)
        role.group.delete()
        messages.success(request, 'Role deleted successfully.')
        return redirect('role_control')


class SchoolUserCreateView(SchoolAccessMixin, View):
    required_permissions = 'auth.add_user'
    template_name = 'core/role_user_form.html'

    def get(self, request):
        return render(request, self.template_name, {'form': SchoolUserForm(school=request.school), 'title': 'Add User'})

    def post(self, request):
        form = SchoolUserForm(request.POST, school=request.school)
        if form.is_valid():
            user = get_user_model().objects.create_user(
                username=form.cleaned_data['username'],
                email=form.cleaned_data['email'],
                password=form.cleaned_data['password'],
                first_name=form.cleaned_data['first_name'],
                last_name=form.cleaned_data['last_name'],
                is_active=form.cleaned_data['is_active'],
                is_staff=False,
            )
            UserProfile.objects.create(user=user, school=request.school, is_school_admin=False)
            user.groups.set([role.group for role in form.cleaned_data['roles']])
            messages.success(request, 'User created successfully.')
            return redirect('role_control')
        return render(request, self.template_name, {'form': form, 'title': 'Add User'})


class SchoolUserUpdateView(SchoolAccessMixin, View):
    required_permissions = 'auth.change_user'
    template_name = 'core/role_user_form.html'

    def get(self, request, pk):
        user = get_object_or_404(_school_user_queryset(request), pk=pk)
        return render(request, self.template_name, {'form': SchoolUserForm(school=request.school, instance=user), 'managed_user': user, 'title': 'Edit User'})

    def post(self, request, pk):
        user = get_object_or_404(_school_user_queryset(request), pk=pk)
        form = SchoolUserForm(request.POST, school=request.school, instance=user)
        if form.is_valid():
            user.username = form.cleaned_data['username']
            user.email = form.cleaned_data['email']
            user.first_name = form.cleaned_data['first_name']
            user.last_name = form.cleaned_data['last_name']
            user.is_active = form.cleaned_data['is_active']
            password = form.cleaned_data['password']
            if password:
                user.set_password(password)
            user.save()
            user.groups.set([role.group for role in form.cleaned_data['roles']])
            messages.success(request, 'User updated successfully.')
            return redirect('role_control')
        return render(request, self.template_name, {'form': form, 'managed_user': user, 'title': 'Edit User'})


def _file_uri(file_field):
    try:
        if not file_field:
            return ''
        if getattr(file_field, 'url', ''):
            return file_field.url
        if Path(file_field.path).exists():
            return Path(file_field.path).as_uri()
    except Exception:
        return ''
    return ''


def _document_preview_response(request, title, html_url):
    return render(
        request,
        'core/document_preview.html',
        {
            'title': title,
            'preview_url': html_url,
        },
    )


def _student_document_url_template(route_name):
    return reverse(route_name, kwargs={'student_id': 0}).replace('/0/', '/__student_id__/')


def _next_serial(school, doc_type):
    # Keep yearly running counters per document type (BONAFIDE, etc.).
    current_year = date.today().year
    counter, _ = DocumentCounter.objects.get_or_create(school=school, doc_type=doc_type, year=current_year)
    counter.last_number += 1
    counter.save(update_fields=['last_number'])
    return f'{doc_type[:3].upper()}-{current_year}-{counter.last_number:04d}'


def _increment_document_count(school, doc_type, amount=1):
    if amount <= 0:
        return
    current_year = date.today().year
    counter, _ = DocumentCounter.objects.get_or_create(school=school, doc_type=doc_type, year=current_year)
    counter.last_number = F('last_number') + amount
    counter.save(update_fields=['last_number'])


def _document_count_totals(school):
    counters = (
        DocumentCounter.objects
        .filter(school=school, doc_type__in=['bonafide', 'id-card', 'nirgam', 'hall-ticket'])
        .values('doc_type')
        .annotate(total=Sum('last_number'))
    )
    totals = {row['doc_type']: row['total'] or 0 for row in counters}
    return {
        'bonafide': totals.get('bonafide', 0),
        'id_card': totals.get('id-card', 0),
        'nirgam': totals.get('nirgam', 0),
        'hall_ticket': totals.get('hall-ticket', 0),
    }


def _academic_year_string(current_date):
    start_year = current_date.year if current_date.month >= 6 else current_date.year - 1
    end_year_short = str(start_year + 1)[-2:]
    return f'{start_year}-{end_year_short}'


ID_CARD_WIDTH = 650
ID_CARD_HEIGHT = 1028


def _existing_path(*candidates):
    for candidate in candidates:
        path = Path(candidate)
        if path.exists():
            return path
    return None


def _load_font(font_name, size):
    font_path = _existing_path(
        Path('C:/Windows/Fonts') / font_name,
        Path('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'),
    )
    if font_path:
        try:
            return ImageFont.truetype(str(font_path), size=size)
        except Exception:
            pass
    return ImageFont.load_default()


def _open_image(path, size=None, opacity=None):
    if not path or not Path(path).exists():
        return None

    image = Image.open(path).convert('RGBA')
    if size:
        image = image.resize(size, Image.Resampling.LANCZOS)
    if opacity is not None:
        alpha = image.getchannel('A')
        alpha = alpha.point(lambda value: int(value * opacity))
        image.putalpha(alpha)
    return image


def _cover_image(path, size):
    if not path or not Path(path).exists():
        return None
    image = Image.open(path).convert('RGB')
    return ImageOps.fit(image, size, Image.Resampling.LANCZOS)


def _student_photo_path(student):
    try:
        if student.photo and Path(student.photo.path).exists():
            return Path(student.photo.path)
    except Exception:
        return None
    return None


def _id_card_assets():
    image_dir = Path(settings.BASE_DIR) / 'static' / 'images'
    return {
        'header': _existing_path(image_dir / 'header.jpeg'),
        'footer': _existing_path(image_dir / 'footer.png'),
        'watermark': _existing_path(
            image_dir / 'School_Logo.png',
            image_dir / 'school_logo.png',
            image_dir / 'school_logo.jpeg',
        ),
    }


def _draw_multiline_text(draw, text, box, font, fill, line_spacing=18):
    x, y, width, _ = box
    text = (text or '').strip()
    if not text:
        return

    words = text.split()
    lines = []
    current = ''
    for word in words:
        candidate = f'{current} {word}'.strip()
        measured_width = draw.textbbox((0, 0), candidate, font=font)[2]
        if current and measured_width > width:
            lines.append(current)
            current = word
        else:
            current = candidate
    if current:
        lines.append(current)

    line_height = draw.textbbox((0, 0), 'Ag', font=font)[3]
    for line in lines:
        draw.text((x, y), line, font=font, fill=fill)
        y += line_height + line_spacing


def _render_id_card_front(student):
    assets = _id_card_assets()
    canvas = Image.new('RGB', (ID_CARD_WIDTH, ID_CARD_HEIGHT), 'white')
    draw = ImageDraw.Draw(canvas)

    watermark = _open_image(assets['watermark'], size=(400, 400), opacity=0.08)
    if watermark:
        canvas.paste(watermark, (125, 636), watermark)

    header = _open_image(assets['header'], size=(ID_CARD_WIDTH, 210))
    if header:
        canvas.paste(header, (0, 0), header)

    photo = _cover_image(_student_photo_path(student), (160, 200))
    if photo:
        canvas.paste(photo, (245, 250))
        draw.rectangle((245, 250, 405, 450), outline='black', width=3)
    else:
        draw.rectangle((245, 250, 405, 450), outline='black', width=3)
        placeholder_font = _load_font('arial.ttf', 22)
        draw.text((280, 336), 'No Photo', font=placeholder_font, fill='black')

    info_font = _load_font('arial.ttf', 32)
    draw.text((40, 525), f"Student's Name : {student.name}", font=info_font, fill='black')
    draw.text((40, 610), f"Father's Name : {student.father_name}", font=info_font, fill='black')
    class_label = f"Class : {student.student_class}"
    if student.section:
        class_label += f' - {student.section}'
    draw.text((40, 695), class_label, font=info_font, fill='black')

    footer = _open_image(assets['footer'], size=(ID_CARD_WIDTH, 220))
    if footer:
        canvas.paste(footer, (0, ID_CARD_HEIGHT - 210), footer)

    return canvas


def _render_id_card_back(student, academic_year):
    assets = _id_card_assets()
    canvas = Image.new('RGB', (ID_CARD_WIDTH, ID_CARD_HEIGHT), 'white')
    draw = ImageDraw.Draw(canvas)

    watermark = _open_image(assets['watermark'], size=(400, 400), opacity=0.08)
    if watermark:
        canvas.paste(watermark, (125, 450), watermark)

    draw.rectangle((0, 0, ID_CARD_WIDTH, 120), fill='#274e13')
    title_font = _load_font('arialbd.ttf', 48)
    draw.text((40, 28), 'Details of Student', font=title_font, fill='white')

    info_font = _load_font('arial.ttf', 40)
    line_y = 200
    line_gap = 120
    fill = '#1a145a'

    draw.text((60, line_y), f'Date of Birth : {student.dob.strftime("%d/%m/%Y")}', font=info_font, fill=fill)
    line_y += line_gap
    draw.text((60, line_y), f'Admission No. : {student.admission_no}', font=info_font, fill=fill)
    line_y += line_gap
    draw.text((60, line_y), f'Contact No. : {student.mobile}', font=info_font, fill=fill)
    line_y += line_gap
    _draw_multiline_text(draw, f'Address : {student.address}', (60, line_y, 530, 220), info_font, fill, line_spacing=12)
    line_y += 220
    draw.text((60, line_y), f'Academic Year : {academic_year}', font=info_font, fill=fill)

    return canvas


def _render_id_card_sheet(student, academic_year):
    front = _render_id_card_front(student)
    back = _render_id_card_back(student, academic_year)
    sheet = Image.new('RGB', (ID_CARD_WIDTH, ID_CARD_HEIGHT * 2), 'white')
    sheet.paste(front, (0, 0))
    sheet.paste(back, (0, ID_CARD_HEIGHT))
    return sheet


def _jpeg_http_response(image, filename):
    buffer = BytesIO()
    image.save(buffer, format='JPEG', quality=92)
    return HttpResponse(
        buffer.getvalue(),
        content_type='image/jpeg',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


def _id_card_payload(request, student_id):
    student = get_object_or_404(_student_queryset(request), id=student_id)
    school = _school_setting(request)
    template_setting = _document_template_settings(request, 'id-card')
    academic_year = _academic_year_string(date.today())
    student_image = _file_uri(student.photo)
    school_logo = (
        _file_uri(template_setting.custom_logo)
        if template_setting and template_setting.custom_logo
        else _file_uri(school.logo) if school else ''
    )
    student_card = {
        'student': student,
        'student_image': student_image,
    }
    context = {
        'student': student,
        'school': school,
        'student_image': student_image,
        'student_cards': [student_card],
        'school_logo': school_logo,
        'school_title': (
            template_setting.school_title
            if template_setting and template_setting.school_title
            else school.school_name if school else 'School Name'
        ),
        'idcard_subtitle': (
            template_setting.school_address
            if template_setting and template_setting.school_address
            else 'Official School Card'
        ),
        'idcard_logo_position': (
            template_setting.logo_position
            if template_setting and template_setting.logo_position in {'left', 'right'}
            else 'left'
        ),
        'template_setting': template_setting,
        'academic_year': academic_year,
        'principal_signature': (
            _file_uri(template_setting.custom_signature)
            if template_setting and template_setting.custom_signature
            else _file_uri(school.principal_signature) if school else ''
        ),
    }
    return {
        'template_name': 'core/pdf/primary id card bulk.html',
        'context': context,
        'filename': f'id-card-{student.admission_no}.pdf',
    }


def _id_card_bulk_payload(request):
    class_filter = request.GET.get('class', '').strip()
    section_filter = request.GET.get('section', '').strip()
    if not class_filter:
        raise Http404('Class is required for bulk ID generation.')

    students = _student_queryset(request).filter(status='active', student_class=class_filter)
    if section_filter:
        students = students.filter(section=section_filter)
    students = students.order_by('student_class', 'section', 'roll_no', 'name')

    if not students.exists():
        raise Http404('No active students found for selected class/section.')

    school = _school_setting(request)
    template_setting = _document_template_settings(request, 'id-card')
    academic_year = _academic_year_string(date.today())
    student_cards = [
        {
            'student': student,
            'student_image': _file_uri(student.photo),
        }
        for student in students
    ]
    # Keep compatibility with template variants expecting `entries`.
    entries = [
        {
            'student': card['student'],
            'student_photo': card['student_image'],
            'seat_number': '',
        }
        for card in student_cards
    ]
    section_suffix = f'-section-{section_filter}' if section_filter else ''

    context = {
        'school': school,
        'school_logo': (
            _file_uri(template_setting.custom_logo)
            if template_setting and template_setting.custom_logo
            else _file_uri(school.logo) if school else ''
        ),
        'school_title': (
            template_setting.school_title
            if template_setting and template_setting.school_title
            else school.school_name if school else 'School Name'
        ),
        'idcard_subtitle': (
            template_setting.school_address
            if template_setting and template_setting.school_address
            else 'Official School Card'
        ),
        'idcard_logo_position': (
            template_setting.logo_position
            if template_setting and template_setting.logo_position in {'left', 'right'}
            else 'left'
        ),
        'template_setting': template_setting,
        'academic_year': academic_year,
        'principal_signature': (
            _file_uri(template_setting.custom_signature)
            if template_setting and template_setting.custom_signature
            else _file_uri(school.principal_signature) if school else ''
        ),
        'selected_class': class_filter,
        'selected_section': section_filter,
        'student_cards': student_cards,
        'entries': entries,
    }
    return {
        'template_name': 'core/pdf/primary id card bulk.html',
        'context': context,
        'filename': f'id-cards-class-{class_filter}{section_suffix}.pdf',
    }


def _bonafide_payload(request, student_id):
    student = get_object_or_404(_student_queryset(request), id=student_id)
    school = _school_setting(request)
    template_setting = _document_template_settings(request, 'bonafide')
    today_value = date.today()
    academic_year = _academic_year_string(today_value)

    try:
        class_number = int(student.student_class)
    except (TypeError, ValueError):
        class_number = None

    if request.GET.get('template') == 'bona':
        template_name = 'core/pdf/bona.html'
    else:
        template_name = 'core/pdf/bona.html' if class_number in (11, 12) else 'core/pdf/bona2.html'
    fixed_header_path = Path(settings.BASE_DIR) / 'templates' / 'core' / 'pdf' / 'img.png'
    bonafide_header_image = fixed_header_path.as_uri() if fixed_header_path.exists() else ''
    fixed_logo_path = Path(settings.BASE_DIR) / 'templates' / 'core' / 'pdf' / 'bonafide_logo.png'

    if template_setting and template_setting.custom_logo:
        bonafide_logo_image = _file_uri(template_setting.custom_logo)
    elif fixed_logo_path.exists():
        bonafide_logo_image = fixed_logo_path.as_uri()
    else:
        bonafide_logo_image = _file_uri(school.logo) if school else ''

    school_title = (
        template_setting.school_title
        if template_setting and template_setting.school_title
        else (school.school_name if school else 'School Name')
    )
    school_address = (
        template_setting.school_address
        if template_setting and template_setting.school_address
        else (school.address if school else '')
    )

    context = {
        'student': student,
        'school': school,
        'template_setting': template_setting,
        'today_date': today_value.strftime('%d/%m/%Y'),
        'academic_year': academic_year,
        'bonafide_header_image': bonafide_header_image,
        'bonafide_logo_image': bonafide_logo_image,
        'bonafide_student_image': _file_uri(student.photo),
        'bonafide_school_title': school_title,
        'bonafide_school_address': school_address,
        'bonafide_title_font_size': template_setting.title_font_size if template_setting else 30,
        'bonafide_address_font_size': template_setting.text_font_size if template_setting else 14,
        'bonafide_logo_position': template_setting.logo_position if template_setting else 'left',
        'bonafide_show_student_photo': (
            template_setting.show_student_photo if template_setting else True
        ),
        'bonafide_border_style': template_setting.border_style if template_setting else 'double',
        'bonafide_border_color': template_setting.border_color if template_setting else '#111111',
        'bonafide_border_width': template_setting.border_width if template_setting else 4,
    }
    return {
        'template_name': template_name,
        'context': context,
        'filename': f'bonafide-{student.admission_no}.pdf',
    }


def _nirgam_payload(request, student_id):
    student = get_object_or_404(_student_queryset(request), id=student_id)
    school = _school_setting(request)
    template_setting = _document_template_settings(request, 'nirgam')
    leaving_date = request.GET.get('leaving_date')
    reason = request.GET.get('reason', '')
    conduct_remarks = request.GET.get('conduct_remarks', '')
    language = request.GET.get('language', 'mr')
    default_titles = {
        'en': (
            'Student Admission Register (As per Govt. Rules)',
            'Student Admission Leaving Register',
        ),
        'mr': (
            'विद्यार्थी प्रवेश रजिस्टर (शासन नियमानुसार)',
            'विद्यार्थी प्रवेश निर्गम नोंदणी',
        ),
    }
    default_title_primary, default_title_secondary = default_titles.get(language, default_titles['mr'])
    saved_title_primary = template_setting.school_title if template_setting and template_setting.school_title else ''
    saved_title_secondary = template_setting.school_address if template_setting and template_setting.school_address else ''

    context = {
        'student': student,
        'school': school,
        'leaving_date': leaving_date,
        'reason': reason,
        'conduct_remarks': conduct_remarks,
        'today': date.today(),
        'school_logo': _file_uri(school.logo) if school else '',
        'principal_signature': _file_uri(school.principal_signature) if school else '',
        'language': language,
        'title_primary': request.GET.get('title_primary') or saved_title_primary or default_title_primary,
        'title_secondary': request.GET.get('title_secondary') or saved_title_secondary or default_title_secondary,
    }
    return {
        'template_name': (
            'core/pdf/nirgam_english.html'
            if language == 'en'
            else 'core/pdf/nirgam_pdf.html'
        ),
        'context': context,
        'filename': f'nirgam-{student.admission_no}.pdf',
    }


def _hallticket_payload(request, student_id):
    student = get_object_or_404(_student_queryset(request), id=student_id)
    school = _school_setting(request)
    exam_id = request.GET.get('exam_id')
    seat_number = request.GET.get('seat_number', '')
    if not exam_id:
        raise Http404('Exam is required')

    exam = get_object_or_404(_exam_queryset(request).prefetch_related('subjects'), id=exam_id)
    template_setting = _document_template_settings(request, 'hall-ticket')
    exam_name_normalized = (exam.name or '').strip().lower()
    is_second_term = any(
        marker in exam_name_normalized
        for marker in ['second term', 'secound term', '2nd term', '2st term']
    )

    context = {
        'student': student,
        'school': school,
        'exam': exam,
        'subjects': exam.subjects.all(),
        'seat_number': seat_number,
        'today': date.today(),
        'school_logo': _file_uri(school.logo) if school else '',
        'school_title': (
            template_setting.school_title
            if template_setting and template_setting.school_title
            else school.school_name if school else 'School Name'
        ),
        'school_address': (
            template_setting.school_address
            if template_setting and template_setting.school_address
            else school.address if school else ''
        ),
        'principal_signature': (
            _file_uri(template_setting.custom_signature)
            if template_setting and template_setting.custom_signature
            else _file_uri(school.principal_signature) if school else ''
        ),
        'student_photo': _file_uri(student.photo),
        'template_setting': template_setting,
    }
    return {
        'template_name': (
            'core/pdf/hall ticket2st.html'
            if is_second_term
            else 'core/pdf/hall ticket.html'
        ),
        'context': context,
        'filename': f'hall-ticket-{student.admission_no}.pdf',
    }


def _hallticket_bulk_payload(request):
    exam_id = request.GET.get('exam_id') or request.GET.get('exam')
    class_filter = request.GET.get('class', '').strip()
    section_filter = request.GET.get('section', '').strip()
    if not exam_id:
        raise Http404('Exam is required')
    if not class_filter:
        raise Http404('Class is required')

    exam = get_object_or_404(_exam_queryset(request).prefetch_related('subjects'), id=exam_id)
    school = _school_setting(request)
    template_setting = _document_template_settings(request, 'hall-ticket')
    students = _student_queryset(request).filter(status='active', student_class=class_filter)
    if section_filter:
        students = students.filter(section=section_filter)
    students = students.order_by('student_class', 'section', 'roll_no', 'name')
    if not students.exists():
        raise Http404('No active students found for selected class/section.')

    entries = []
    for idx, student in enumerate(students, start=1):
        section_value = section_filter or (student.section or '')
        seat_number = f'{class_filter}{section_value}-{idx:03d}' if section_value else f'{class_filter}-{idx:03d}'
        entries.append(
            {
                'student': student,
                'seat_number': seat_number,
                'student_photo': _file_uri(student.photo),
            }
        )

    exam_name_normalized = (exam.name or '').strip().lower()
    is_second_term = any(
        marker in exam_name_normalized
        for marker in ['second term', 'secound term', '2nd term', '2st term']
    )

    context = {
        'school': school,
        'exam': exam,
        'subjects': exam.subjects.all(),
        'entries': entries,
        'selected_class': class_filter,
        'selected_section': section_filter,
        'today': date.today(),
        'school_logo': _file_uri(school.logo) if school else '',
        'school_title': (
            template_setting.school_title
            if template_setting and template_setting.school_title
            else school.school_name if school else 'School Name'
        ),
        'school_address': (
            template_setting.school_address
            if template_setting and template_setting.school_address
            else school.address if school else ''
        ),
        'principal_signature': (
            _file_uri(template_setting.custom_signature)
            if template_setting and template_setting.custom_signature
            else _file_uri(school.principal_signature) if school else ''
        ),
        'template_setting': template_setting,
    }
    return {
        'template_name': (
            'core/pdf/hall ticket2st bulk.html'
            if is_second_term
            else 'core/pdf/hall ticket bulk.html'
        ),
        'context': context,
        'filename': f'hall-tickets-{class_filter}-{section_filter or "all"}.pdf',
    }


class IDCardFormView(SchoolScopedModelFormMixin, FormView):
    template_name = 'core/document_form.html'
    form_class = StudentSelectForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Generate ID Card'
        context['submit_label'] = 'Preview ID Card'
        context['secondary_url'] = reverse('idcard_bulk_form')
        context['secondary_label'] = 'Class Wise Bulk ID Cards'
        context['preview_url_template'] = _student_document_url_template('idcard_html')
        context['download_url_template'] = _student_document_url_template('idcard_jpg')
        return context

    def form_valid(self, form):
        student = form.cleaned_data['student']
        _increment_document_count(self.request.school, 'id-card')
        return redirect('idcard_preview', student_id=student.id)


class IDCardBulkFormView(SchoolAccessMixin, FormView):
    template_name = 'core/document_form.html'
    form_class = IDCardBulkForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Generate ID Cards (Class Wise)'
        context['submit_label'] = 'Preview Bulk ID Cards'
        return context

    def form_valid(self, form):
        class_filter = form.cleaned_data['student_class']
        section_filter = form.cleaned_data.get('section', '')
        students = _student_queryset(self.request).filter(status='active', student_class=class_filter)
        if section_filter:
            students = students.filter(section=section_filter)
        _increment_document_count(self.request.school, 'id-card', students.count())
        query_string = urlencode(
            {
                'class': class_filter,
                'section': section_filter,
            }
        )
        return redirect(f"{reverse_lazy('idcard_bulk_preview')}?{query_string}")


class IDCardPreviewView(SchoolAccessMixin, View):
    def get(self, request, student_id):
        return render(
            request,
            'core/document_preview_idcard.html',
            {
                'title': 'ID Card Preview',
                'preview_url': reverse('idcard_html', kwargs={'student_id': student_id}),
                'jpg_url': reverse('idcard_jpg', kwargs={'student_id': student_id}),
            },
        )


class IDCardBulkPreviewView(SchoolAccessMixin, View):
    def get(self, request):
        class_filter = request.GET.get('class', '').strip()
        section_filter = request.GET.get('section', '').strip()
        if not class_filter:
            messages.error(request, 'Select a class first to generate bulk ID cards.')
            return redirect('student_list')

        query_string = urlencode({'class': class_filter, 'section': section_filter})
        return render(
            request,
            'core/document_preview_idcard.html',
            {
                'title': 'Bulk ID Card Preview',
                'preview_url': f"{reverse('idcard_bulk_html')}?{query_string}",
                'jpg_url': f"{reverse('idcard_bulk_jpg')}?{query_string}",
            },
        )


class IDCardHTMLView(SchoolAccessMixin, View):
    def get(self, request, student_id):
        if request.GET.get('generated') == '1':
            _increment_document_count(request.school, 'id-card')
        payload = _id_card_payload(request, student_id)
        return render(request, payload['template_name'], payload['context'])


class IDCardPDFView(SchoolAccessMixin, View):
    def get(self, request, student_id):
        return redirect('idcard_html', student_id=student_id)


class IDCardJPGView(SchoolAccessMixin, View):
    def get(self, request, student_id):
        student = get_object_or_404(_student_queryset(request), id=student_id)
        blocked_response = _authorize_download(request, reason='id-card-jpg')
        if blocked_response is not None:
            return blocked_response
        academic_year = _academic_year_string(date.today())
        image = _render_id_card_sheet(student, academic_year)
        return _jpeg_http_response(image, f'id-card-{student.admission_no}.jpg')


class IDCardBulkHTMLView(SchoolAccessMixin, View):
    def get(self, request):
        payload = _id_card_bulk_payload(request)
        return render(request, payload['template_name'], payload['context'])


class IDCardBulkPDFView(SchoolAccessMixin, View):
    def get(self, request):
        query_string = request.GET.urlencode()
        html_url = reverse('idcard_bulk_html')
        if query_string:
            html_url = f'{html_url}?{query_string}'
        return redirect(html_url)


class IDCardBulkJPGView(SchoolAccessMixin, View):
    def get(self, request):
        class_filter = request.GET.get('class', '').strip()
        section_filter = request.GET.get('section', '').strip()
        if not class_filter:
            raise Http404('Class is required for bulk ID generation.')

        students = _student_queryset(request).filter(status='active', student_class=class_filter)
        if section_filter:
            students = students.filter(section=section_filter)
        students = students.order_by('student_class', 'section', 'roll_no', 'name')
        if not students.exists():
            raise Http404('No active students found for selected class/section.')

        blocked_response = _authorize_download(request, reason='bulk-id-card-jpg')
        if blocked_response is not None:
            return blocked_response

        academic_year = _academic_year_string(date.today())
        buffer = BytesIO()
        with ZipFile(buffer, 'w', compression=ZIP_DEFLATED) as zip_file:
            for student in students:
                image = _render_id_card_sheet(student, academic_year)
                image_buffer = BytesIO()
                image.save(image_buffer, format='JPEG', quality=92)
                zip_file.writestr(f'id-card-{student.admission_no}.jpg', image_buffer.getvalue())

        section_suffix = f'-section-{section_filter}' if section_filter else ''
        response = HttpResponse(buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = (
            f'attachment; filename="id-cards-class-{class_filter}{section_suffix}.zip"'
        )
        return response


class BonafideFormView(SchoolScopedModelFormMixin, FormView):
    template_name = 'core/document_form.html'
    form_class = BonafideForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Generate Bonafide Certificate'
        context['submit_label'] = 'Preview Bonafide'
        context['preview_url_template'] = _student_document_url_template('bonafide_html')
        return context

    def form_valid(self, form):
        student = form.cleaned_data['student']
        _increment_document_count(self.request.school, 'bonafide')
        return redirect('bonafide_preview', student_id=student.id)


class BonafidePreviewView(SchoolAccessMixin, View):
    def get(self, request, student_id):
        return render(
            request,
            'core/document_preview_a4.html',
            {
                'title': 'Bonafide Preview',
                'preview_url': reverse('bonafide_html', kwargs={'student_id': student_id}),
            },
        )


class BonafideHTMLView(SchoolAccessMixin, View):
    def get(self, request, student_id):
        if request.GET.get('generated') == '1':
            _increment_document_count(request.school, 'bonafide')
        payload = _bonafide_payload(request, student_id)
        return render(request, payload['template_name'], payload['context'])


class BonafidePDFView(SchoolAccessMixin, View):
    def get(self, request, student_id):
        return redirect('bonafide_html', student_id=student_id)


class NirgamFormView(SchoolScopedModelFormMixin, FormView):
    template_name = 'core/document_form.html'
    form_class = NirgamForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Generate Nirgam Certificate'
        context['submit_label'] = 'Preview Nirgam'
        context['preview_url_template'] = _student_document_url_template('nirgam_html')
        return context

    def form_valid(self, form):
        student = form.cleaned_data['student']
        _increment_document_count(self.request.school, 'nirgam')
        leaving_date = form.cleaned_data['leaving_date']
        language = form.cleaned_data['language']
        reason = form.cleaned_data['reason']
        conduct_remarks = form.cleaned_data['conduct_remarks']
        query_string = urlencode(
            {
                'leaving_date': leaving_date,
                'language': language,
                'reason': reason,
                'conduct_remarks': conduct_remarks,
            }
        )
        url = f"{reverse_lazy('nirgam_preview', kwargs={'student_id': student.id})}?{query_string}"
        return redirect(url)


class NirgamPreviewView(SchoolAccessMixin, View):
    def get(self, request, student_id):
        query = request.GET.urlencode()
        suffix = f'?{query}' if query else ''
        return render(
            request,
            'core/document_preview_nirgam.html',
            {
                'title': 'Nirgam Preview',
                'preview_url': f"{reverse('nirgam_html', kwargs={'student_id': student_id})}{suffix}",
            },
        )


class NirgamHTMLView(SchoolAccessMixin, View):
    def get(self, request, student_id):
        if request.GET.get('generated') == '1':
            _increment_document_count(request.school, 'nirgam')
        payload = _nirgam_payload(request, student_id)
        return render(request, payload['template_name'], payload['context'])


class NirgamPDFView(SchoolAccessMixin, View):
    def get(self, request, student_id):
        query_string = request.GET.urlencode()
        html_url = reverse('nirgam_html', kwargs={'student_id': student_id})
        if query_string:
            html_url = f'{html_url}?{query_string}'
        return redirect(html_url)


class HallTicketFormView(SchoolScopedModelFormMixin, FormView):
    template_name = 'core/document_form.html'
    form_class = HallTicketForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Generate Hall Ticket'
        context['submit_label'] = 'Preview Hall Ticket'
        context['secondary_url'] = reverse('hallticket_bulk_form')
        context['secondary_label'] = 'Class Wise Bulk Hall Tickets'
        context['no_exams'] = not _exam_queryset(self.request).exists()
        context['preview_url_template'] = _student_document_url_template('hallticket_html')
        return context

    def form_valid(self, form):
        student = form.cleaned_data['student']
        exam = form.cleaned_data['exam']
        seat_number = form.cleaned_data['seat_number']
        _increment_document_count(self.request.school, 'hall-ticket')
        query_string = urlencode({'exam_id': exam.id, 'seat_number': seat_number})
        url = f"{reverse_lazy('hallticket_preview', kwargs={'student_id': student.id})}?{query_string}"
        return redirect(url)


class HallTicketBulkFormView(SchoolScopedModelFormMixin, FormView):
    template_name = 'core/document_form.html'
    form_class = HallTicketBulkForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Generate Hall Tickets (Class Wise)'
        context['submit_label'] = 'Preview Bulk Hall Tickets'
        context['no_exams'] = not _exam_queryset(self.request).exists()
        return context

    def form_valid(self, form):
        exam = form.cleaned_data['exam']
        class_filter = form.cleaned_data['student_class']
        section_filter = form.cleaned_data.get('section', '')
        students = _student_queryset(self.request).filter(status='active', student_class=class_filter)
        if section_filter:
            students = students.filter(section=section_filter)
        _increment_document_count(self.request.school, 'hall-ticket', students.count())
        query_string = urlencode(
            {
                'exam_id': exam.id,
                'class': class_filter,
                'section': section_filter,
            }
        )
        return redirect(f"{reverse_lazy('hallticket_bulk_preview')}?{query_string}")


class HallTicketPreviewView(SchoolAccessMixin, View):
    def get(self, request, student_id):
        query = request.GET.urlencode()
        suffix = f'?{query}' if query else ''
        return _document_preview_response(
            request,
            title='Hall Ticket Preview',
            html_url=f"{reverse('hallticket_html', kwargs={'student_id': student_id})}{suffix}",
        )


class HallTicketHTMLView(SchoolAccessMixin, View):
    def get(self, request, student_id):
        if request.GET.get('generated') == '1':
            _increment_document_count(request.school, 'hall-ticket')
        payload = _hallticket_payload(request, student_id)
        return render(request, payload['template_name'], payload['context'])


class HallTicketPDFView(SchoolAccessMixin, View):
    def get(self, request, student_id):
        query_string = request.GET.urlencode()
        html_url = reverse('hallticket_html', kwargs={'student_id': student_id})
        if query_string:
            html_url = f'{html_url}?{query_string}'
        return redirect(html_url)


class HallTicketBulkPreviewView(SchoolAccessMixin, View):
    def get(self, request):
        query = request.GET.urlencode()
        if not query:
            messages.error(request, 'Select exam and class for bulk hall tickets.')
            return redirect('hallticket_bulk_form')
        return render(
            request,
            'core/document_preview_a4.html',
            {
                'title': 'Bulk Hall Ticket Preview',
                'preview_url': f"{reverse('hallticket_bulk_html')}?{query}",
            },
        )


class HallTicketBulkHTMLView(SchoolAccessMixin, View):
    def get(self, request):
        payload = _hallticket_bulk_payload(request)
        return render(request, payload['template_name'], payload['context'])


class HallTicketBulkPDFView(SchoolAccessMixin, View):
    def get(self, request):
        query_string = request.GET.urlencode()
        html_url = reverse('hallticket_bulk_html')
        if query_string:
            html_url = f'{html_url}?{query_string}'
        return redirect(html_url)


class DatabaseBackupView(SchoolAccessMixin, View):
    def get(self, request):
        db_path = Path(settings.DATABASES['default']['NAME'])
        if not db_path.exists():
            raise Http404('Database file not found')

        blocked_response = _authorize_download(request, reason='database-backup')
        if blocked_response is not None:
            return blocked_response

        messages.success(request, 'Database backup downloaded successfully.')
        return FileResponse(open(db_path, 'rb'), as_attachment=True, filename='db.sqlite3')


class SubscriptionView(SchoolAccessMixin, TemplateView):
    template_name = 'core/subscription.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        subscription = _school_subscription(self.request.school)
        context['subscription'] = subscription
        context['next_url'] = self.request.GET.get('next', '').strip()
        context['reason'] = self.request.GET.get('reason', '').strip()
        context['razorpay_key_id'] = settings.RAZORPAY_KEY_ID
        context['razorpay_enabled'] = _razorpay_is_configured()
        context['payment_create_url'] = reverse('subscription_payment_create')
        context['payment_verify_url'] = reverse('subscription_payment_verify')
        context['company_name'] = settings.RAZORPAY_COMPANY_NAME
        return context


class CreateSubscriptionPaymentView(SchoolAccessMixin, View):
    def post(self, request):
        if not _razorpay_is_configured():
            return JsonResponse({'ok': False, 'message': 'Razorpay keys are not configured.'}, status=400)

        subscription = _school_subscription(request.school)
        if subscription.annual_plan_is_active:
            return JsonResponse({'ok': False, 'message': 'Annual plan is already active.'}, status=400)

        receipt = f'sub_{request.school.id}_{uuid4().hex[:18]}'
        notes = {
            'school_id': str(request.school.id),
            'subscription_id': str(subscription.id),
            'plan': 'annual',
        }

        try:
            order = _razorpay_client().order.create(
                {
                    'amount': subscription.annual_price * 100,
                    'currency': settings.RAZORPAY_CURRENCY,
                    'receipt': receipt,
                    'notes': notes,
                }
            )
        except Exception as exc:
            logger.exception('Failed to create Razorpay order.')
            return JsonResponse({'ok': False, 'message': f'Could not create Razorpay order: {exc}'}, status=502)

        payment = SubscriptionPayment.objects.create(
            school=request.school,
            subscription=subscription,
            receipt=receipt,
            order_id=order['id'],
            amount=order['amount'],
            currency=order['currency'],
            notes_json=notes,
        )

        return JsonResponse(
            {
                'ok': True,
                'order_id': payment.order_id,
                'amount': payment.amount,
                'currency': payment.currency,
                'key': settings.RAZORPAY_KEY_ID,
                'name': settings.RAZORPAY_COMPANY_NAME,
                'description': 'EduFlow 1 Year Plan',
            }
        )


class VerifySubscriptionPaymentView(SchoolAccessMixin, View):
    def post(self, request):
        if not _razorpay_is_configured():
            return JsonResponse({'ok': False, 'message': 'Razorpay keys are not configured.'}, status=400)

        razorpay_payment_id = request.POST.get('razorpay_payment_id', '').strip()
        razorpay_order_id = request.POST.get('razorpay_order_id', '').strip()
        razorpay_signature = request.POST.get('razorpay_signature', '').strip()
        next_url = request.POST.get('next', '').strip()

        if not all([razorpay_payment_id, razorpay_order_id, razorpay_signature]):
            messages.error(request, 'Payment response from Razorpay is incomplete.')
            return redirect(next_url or 'subscription')

        payment = get_object_or_404(
            SubscriptionPayment.objects.select_related('subscription'),
            school=request.school,
            order_id=razorpay_order_id,
        )

        try:
            _razorpay_client().utility.verify_payment_signature(
                {
                    'razorpay_order_id': razorpay_order_id,
                    'razorpay_payment_id': razorpay_payment_id,
                    'razorpay_signature': razorpay_signature,
                }
            )
        except Exception:
            payment.payment_id = razorpay_payment_id
            payment.signature = razorpay_signature
            payment.status = SubscriptionPayment.STATUS_FAILED
            payment.save(update_fields=['payment_id', 'signature', 'status', 'updated_at'])
            messages.error(request, 'Razorpay payment verification failed.')
            return redirect(next_url or 'subscription')

        payment.payment_id = razorpay_payment_id
        payment.signature = razorpay_signature
        payment.status = SubscriptionPayment.STATUS_VERIFIED
        payment.save(update_fields=['payment_id', 'signature', 'status', 'updated_at'])

        subscription = payment.subscription
        subscription.activate_annual_plan()
        subscription.save(update_fields=['annual_plan_started_on', 'annual_plan_expires_on', 'updated_at'])

        messages.success(
            request,
            f'Payment verified. Annual plan is active until {subscription.annual_plan_expires_on}.',
        )
        return redirect(next_url or 'subscription')


class TemplateCustomizerHomeView(SchoolAccessMixin, TemplateView):
    template_name = 'core/template_customizer_home.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['template_options'] = [
            {
                'title': 'ID Card',
                'slug': 'id-card',
                'description': 'Customize school branding and student card layout.',
                'icon': 'ti-id',
            },
            {
                'title': 'Bonafide',
                'slug': 'bonafide',
                'description': 'Adjust certificate layout, visibility, and signature blocks.',
                'icon': 'ti-file-certificate',
            },
            {
                'title': 'Hall Ticket',
                'slug': 'hall-ticket',
                'description': 'Control student info, subjects, signature, and instructions.',
                'icon': 'ti-ticket',
            },
            {
                'title': 'Report Card',
                'slug': 'report-card',
                'description': 'Customize marks table, grading summary, colors, and signature area.',
                'icon': 'ti-report-analytics',
            },
            {
                'title': 'Nirgam',
                'slug': 'nirgam',
                'description': 'Customize school leaving certificate sections and branding.',
                'icon': 'ti-file-description',
            },
        ]
        return context


class DocumentTemplateCustomizerView(SchoolAccessMixin, TemplateView):
    template_name = 'core/document_template_customizer.html'

    DOCUMENT_CONFIGS = {
        'id-card': {
            'title': 'ID Card',
            'subtitle': 'Student identity card preview with editable school branding.',
            'header_subtitle': 'Identity Card | Academic Session 2026-27',
            'student_section_title': 'Student Info',
            'details_section_title': 'Card Details',
            'details_rows': [
                ('ID Valid Till', '31 Mar 2027'),
                ('Transport', 'School Bus - Route 4'),
                ('Emergency Contact', '+91 98765 43210'),
            ],
            'show_subjects': True,
            'show_instructions': False,
            'instructions': [],
        },
        'bonafide': {
            'title': 'Bonafide',
            'subtitle': 'Certificate preview with editable school name and layout controls.',
            'header_subtitle': 'Bonafide Certificate',
            'student_section_title': 'Student Info',
            'details_section_title': 'Certificate Details',
            'details_rows': [
                ('Purpose', 'For bank account opening'),
                ('Academic Year', '2026-27'),
                ('Issue Date', '20 Apr 2026'),
            ],
            'show_subjects': True,
            'show_instructions': False,
            'instructions': [],
        },
        'hall-ticket': {
            'title': 'Hall Ticket',
            'subtitle': 'Examination hall ticket with editable school branding and section toggles.',
            'header_subtitle': 'Hall Ticket | Academic Session 2026-27',
            'student_section_title': 'Student Info',
            'details_section_title': 'Subjects',
            'details_rows': [
                ('English', '12 Apr 2026 | 10:00 AM'),
                ('Mathematics', '14 Apr 2026 | 10:00 AM'),
                ('Science', '16 Apr 2026 | 10:00 AM'),
            ],
            'show_subjects': True,
            'show_instructions': True,
            'instructions': [
                'Carry this hall ticket to every examination.',
                'Reach the exam center at least 30 minutes before the start time.',
                'Use only approved stationery and follow invigilator instructions.',
            ],
        },
        'nirgam': {
            'title': 'Nirgam',
            'subtitle': 'School leaving certificate preview with editable sections.',
            'header_subtitle': 'Nirgam / School Leaving Certificate',
            'student_section_title': 'Student Info',
            'details_section_title': 'Leaving Details',
            'details_rows': [
                ('Leaving Date', '30 Apr 2026'),
                ('Reason', 'Transfer to another school'),
                ('Conduct', 'Good'),
            ],
            'show_subjects': True,
            'show_instructions': True,
            'instructions': [
                'Verify student details before issuing the certificate.',
                'Principal signature and school seal are required.',
            ],
        },
        'report-card': {
            'title': 'Report Card',
            'subtitle': 'Combined first term and second term report with editable branding, totals, and signature area.',
            'header_subtitle': 'Progress Report | First Term and Second Term',
            'student_section_title': 'Student Details',
            'details_section_title': 'Marks Summary',
            'details_rows': [
                ('English', '82 / 100'),
                ('Mathematics', '91 / 100'),
                ('Science', '88 / 100'),
                ('Social Studies', '79 / 100'),
            ],
            'show_subjects': True,
            'show_instructions': True,
            'instructions': [
                'Total Marks: 340 / 400',
                'Percentage: 85%',
                'Result: Pass',
            ],
        },
    }

    def post(self, request, *args, **kwargs):
        document_type = self.kwargs['document_type']
        if document_type == 'id-card':
            existing_setting = DocumentTemplateSetting.objects.filter(
                school=request.school,
                document_type='id-card',
            ).first()
            logo_position = request.POST.get('logo_position', 'left')
            if logo_position not in {'left', 'right'}:
                logo_position = existing_setting.logo_position if existing_setting else 'left'
            if logo_position not in {'left', 'right'}:
                logo_position = 'left'
            template_setting, _ = DocumentTemplateSetting.objects.update_or_create(
                school=request.school,
                document_type='id-card',
                defaults={
                    'school_title': request.POST.get('school_title', '').strip()[:255],
                    'school_address': request.POST.get('school_address', '').strip(),
                    'accent_color': _clean_hex_color(
                        request.POST.get('accent_color'),
                        existing_setting.accent_color if existing_setting else '#274e13',
                    ),
                    'background_color': _clean_hex_color(
                        request.POST.get('background_color'),
                        existing_setting.background_color if existing_setting else '#274e13',
                    ),
                    'text_color': _clean_hex_color(
                        request.POST.get('text_color'),
                        existing_setting.text_color if existing_setting else '#000080',
                    ),
                    'border_color': _clean_hex_color(
                        request.POST.get('border_color'),
                        existing_setting.border_color if existing_setting else '#000000',
                    ),
                    'title_font_size': _clamp_int(
                        request.POST.get('title_font_size'),
                        existing_setting.title_font_size if existing_setting else 38,
                        16,
                        60,
                    ),
                    'text_font_size': _clamp_int(
                        request.POST.get('text_font_size'),
                        existing_setting.text_font_size if existing_setting else 22,
                        10,
                        40,
                    ),
                    'logo_position': logo_position,
                },
            )
            custom_signature = request.FILES.get('custom_signature')
            if custom_signature:
                template_setting.custom_signature = custom_signature
                template_setting.save(update_fields=['custom_signature'])
            custom_logo = request.FILES.get('custom_logo')
            if custom_logo:
                template_setting.custom_logo = custom_logo
                template_setting.save(update_fields=['custom_logo'])
            messages.success(request, 'ID Card template settings saved successfully.')
        elif document_type == 'bonafide':
            existing_setting = DocumentTemplateSetting.objects.filter(
                school=request.school,
                document_type='bonafide',
            ).first()
            logo_position = request.POST.get('logo_position', 'left')
            if logo_position not in {'left', 'right', 'none'}:
                logo_position = existing_setting.logo_position if existing_setting else 'left'
            border_style = request.POST.get('border_style', 'double')
            if border_style not in {'double', 'solid', 'dashed', 'dotted', 'none'}:
                border_style = existing_setting.border_style if existing_setting else 'double'

            template_setting, _ = DocumentTemplateSetting.objects.update_or_create(
                school=request.school,
                document_type='bonafide',
                defaults={
                    'school_title': request.POST.get('school_title', '').strip()[:255],
                    'school_address': request.POST.get('school_address', '').strip(),
                    'title_font_size': _clamp_int(
                        request.POST.get('title_font_size'),
                        existing_setting.title_font_size if existing_setting else 30,
                        18,
                        56,
                    ),
                    'text_font_size': _clamp_int(
                        request.POST.get('text_font_size'),
                        existing_setting.text_font_size if existing_setting else 14,
                        10,
                        28,
                    ),
                    'logo_position': logo_position,
                    'show_student_photo': request.POST.get('show_student_photo') == 'yes',
                    'border_style': border_style,
                    'border_color': _clean_hex_color(
                        request.POST.get('border_color'),
                        existing_setting.border_color if existing_setting else '#111111',
                    ),
                    'border_width': _clamp_int(
                        request.POST.get('border_width'),
                        existing_setting.border_width if existing_setting else 4,
                        1,
                        12,
                    ),
                },
            )
            custom_logo = request.FILES.get('custom_logo')
            if custom_logo:
                template_setting.custom_logo = custom_logo
                template_setting.save(update_fields=['custom_logo'])
            messages.success(request, 'Bonafide template settings saved successfully.')
        elif document_type == 'hall-ticket':
            existing_setting = DocumentTemplateSetting.objects.filter(
                school=request.school,
                document_type='hall-ticket',
            ).first()
            template_setting, _ = DocumentTemplateSetting.objects.update_or_create(
                school=request.school,
                document_type='hall-ticket',
                defaults={
                    'accent_color': request.POST.get(
                        'accent_color',
                        existing_setting.accent_color if existing_setting else '#cb0804',
                    )[:7],
                    'border_color': request.POST.get(
                        'border_color',
                        existing_setting.border_color if existing_setting else '#cb0804',
                    )[:7],
                    'background_color': request.POST.get('background_color', '#f5fcc8')[:7],
                    'text_color': request.POST.get('text_color', '#000080')[:7],
                    'school_title': request.POST.get('school_title', '').strip()[:255],
                    'school_address': request.POST.get('school_address', '').strip(),
                    'title_font_size': _clamp_int(
                        request.POST.get('title_font_size'),
                        existing_setting.title_font_size if existing_setting else 38,
                        20,
                        72,
                    ),
                    'text_font_size': _clamp_int(
                        request.POST.get('text_font_size'),
                        existing_setting.text_font_size if existing_setting else 18,
                        12,
                        36,
                    ),
                    'border_style': (
                        request.POST.get('border_style')
                        if request.POST.get('border_style') in {'accent', 'none'}
                        else (existing_setting.border_style if existing_setting else 'accent')
                    ),
                    'logo_position': request.POST.get('logo_position', 'both'),
                    'text_position': request.POST.get('text_position', 'center'),
                    'image_position': request.POST.get('image_position', 'right'),
                },
            )
            custom_logo = request.FILES.get('custom_logo')
            if custom_logo:
                template_setting.custom_logo = custom_logo
                template_setting.save(update_fields=['custom_logo'])
            custom_signature = request.FILES.get('custom_signature')
            if custom_signature:
                template_setting.custom_signature = custom_signature
                template_setting.save(update_fields=['custom_signature'])
            messages.success(request, 'Hall Ticket template settings saved successfully.')
        elif document_type == 'nirgam':
            DocumentTemplateSetting.objects.update_or_create(
                school=request.school,
                document_type='nirgam',
                defaults={
                    'school_title': request.POST.get('school_title', '').strip()[:255],
                    'school_address': request.POST.get('school_address', '').strip(),
                },
            )
            messages.success(request, 'Nirgam template settings saved successfully.')
        elif document_type == 'report-card':
            existing_setting = DocumentTemplateSetting.objects.filter(
                school=request.school,
                document_type='report-card',
            ).first()
            template_setting, _ = DocumentTemplateSetting.objects.update_or_create(
                school=request.school,
                document_type='report-card',
                defaults={
                    'school_title': request.POST.get('school_title', '').strip()[:255],
                    'school_address': request.POST.get('school_address', '').strip(),
                    'accent_color': _clean_hex_color(
                        request.POST.get('accent_color'),
                        existing_setting.accent_color if existing_setting else '#0f766e',
                    ),
                    'border_color': _clean_hex_color(
                        request.POST.get('border_color'),
                        existing_setting.border_color if existing_setting else '#0f766e',
                    ),
                    'background_color': _clean_hex_color(request.POST.get('background_color'), '#ffffff'),
                    'text_color': _clean_hex_color(request.POST.get('text_color'), '#111827'),
                    'title_font_size': _clamp_int(
                        request.POST.get('title_font_size'),
                        existing_setting.title_font_size if existing_setting else 30,
                        18,
                        56,
                    ),
                    'text_font_size': _clamp_int(
                        request.POST.get('text_font_size'),
                        existing_setting.text_font_size if existing_setting else 14,
                        10,
                        28,
                    ),
                    'border_style': (
                        request.POST.get('border_style')
                        if request.POST.get('border_style') in {'solid', 'double', 'dashed', 'dotted', 'none'}
                        else (existing_setting.border_style if existing_setting else 'solid')
                    ),
                },
            )
            custom_logo = request.FILES.get('custom_logo')
            if custom_logo:
                template_setting.custom_logo = custom_logo
                template_setting.save(update_fields=['custom_logo'])
            custom_signature = request.FILES.get('custom_signature')
            if custom_signature:
                template_setting.custom_signature = custom_signature
                template_setting.save(update_fields=['custom_signature'])
            messages.success(request, 'Report Card template settings saved successfully.')
        else:
            messages.info(request, 'Save is currently enabled for Bonafide and Hall Ticket customization.')
        return redirect('document_template_customizer', document_type=document_type)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        document_type = self.kwargs['document_type']
        config = self.DOCUMENT_CONFIGS.get(document_type)
        if config is None:
            raise Http404('Template type not found.')
        saved_setting = _document_template_settings(self.request, document_type)
        preview_school = _school_setting(self.request)
        id_card_preview_student = None
        id_card_preview_url = ''
        if document_type == 'id-card':
            id_card_preview_student = (
                _student_queryset(self.request)
                .filter(status='active')
                .order_by('student_class', 'section', 'roll_no', 'name')
                .first()
            )
            if id_card_preview_student is not None:
                query_string = urlencode(
                    {
                        'class': id_card_preview_student.student_class,
                        'section': id_card_preview_student.section,
                    }
                )
                id_card_preview_url = f"{reverse('idcard_bulk_html')}?{query_string}"
        bonafide_preview_student = None
        bonafide_preview_url = ''
        if document_type == 'bonafide':
            bonafide_preview_student = (
                _student_queryset(self.request)
                .filter(status='active')
                .order_by('student_class', 'section', 'roll_no', 'name')
                .first()
            )
            if bonafide_preview_student is not None:
                bonafide_preview_url = (
                    f"{reverse('bonafide_html', kwargs={'student_id': bonafide_preview_student.id})}"
                    "?template=bona"
                )
        nirgam_preview_student = None
        nirgam_preview_urls = {}
        if document_type == 'nirgam':
            nirgam_preview_student = (
                _student_queryset(self.request)
                .filter(status='active')
                .order_by('student_class', 'section', 'roll_no', 'name')
                .first()
            )
            if nirgam_preview_student is not None:
                base_query = {
                    'leaving_date': date.today().isoformat(),
                    'reason': 'Transfer to another school',
                    'conduct_remarks': 'Good',
                }
                for language in ('mr', 'en'):
                    query_string = urlencode({**base_query, 'language': language})
                    nirgam_preview_urls[language] = (
                        f"{reverse('nirgam_html', kwargs={'student_id': nirgam_preview_student.id})}"
                        f"?{query_string}"
                    )
        context.update(
            {
                'document_type': document_type,
                'document_config': config,
                'saved_setting': saved_setting,
                'preview_school': preview_school,
                'id_card_preview_student': id_card_preview_student,
                'id_card_preview_url': id_card_preview_url,
                'bonafide_preview_student': bonafide_preview_student,
                'bonafide_preview_url': bonafide_preview_url,
                'nirgam_preview_student': nirgam_preview_student,
                'nirgam_preview_urls': nirgam_preview_urls,
            }
        )
        return context

